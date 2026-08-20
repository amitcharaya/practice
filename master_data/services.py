from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from django.db import transaction
from .models import Jail, SHG, VegetableMaster
from .serializers import (
    JailBulkUploadSerializer,
    SHGBulkUploadSerializer,
    VegetableBulkUploadSerializer,
)


JAIL_HEADERS = [
    "name",
    "location",
    "is_active",
]

SHG_HEADERS = [
    "name",
    "jail",
    "contact_person",
    "is_active",
]

VEGETABLE_HEADERS = [
    "item_name",
    "unit",
    "punjabi_name",
    "category",
    "rate",
    "is_active",
]

def load_excel_file(uploaded_file):
    try:
        workbook = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True
        )
    except Exception:
        raise ValidationError(
            "Invalid Excel file. Please upload a valid .xlsx file."
        )

    return workbook

def get_active_sheet(workbook):
    return workbook.active

def validate_headers(worksheet, expected_headers):

    actual_headers = [
        cell.value
        for cell in worksheet[1]
    ]

    actual_headers = [
        str(header).strip()
        if header is not None
        else ""
        for header in actual_headers
    ]

    if actual_headers != expected_headers:

        missing_headers = [
            header
            for header in expected_headers
            if header not in actual_headers
        ]

        extra_headers = [
            header
            for header in actual_headers
            if header and header not in expected_headers
        ]

        errors = {}

        if missing_headers:
            errors["missing_columns"] = missing_headers

        if extra_headers:
            errors["unexpected_columns"] = extra_headers

        raise ValidationError(errors)

    return actual_headers

def row_to_dict(headers, row):

    return {
        header: value
        for header, value in zip(headers, row)
    }
def is_empty_row(row):
    return all(
        value is None or str(value).strip() == ""
        for value in row
    )

def normalize_boolean(value):

    if isinstance(value, bool):
        return value

    if value is None:
        return False

    value = str(value).strip().lower()

    if value in ["true", "1", "yes"]:
        return True

    if value in ["false", "0", "no"]:
        return False

    raise ValueError(
        f"Invalid boolean value: {value}"
    )

def bulk_upload_jails(uploaded_file):

    workbook = load_excel_file(uploaded_file)
    worksheet = get_active_sheet(workbook)

    headers = validate_headers(
        worksheet,
        JAIL_HEADERS
    )

    records = []
    errors = []

    existing_names = set(
        Jail.objects.values_list(
            "name",
            flat=True
        )
    )

    uploaded_names = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if is_empty_row(row):
            continue

        data = row_to_dict(headers, row)

        try:
            data["is_active"] = normalize_boolean(
                data.get("is_active")
            )

            name = str(
                data.get("name") or ""
            ).strip()

            data["name"] = name

            if not name:
                raise ValueError(
                    "Jail name is required."
                )

            if name in existing_names:
                raise ValueError(
                    f"Jail '{name}' already exists."
                )

            if name in uploaded_names:
                raise ValueError(
                    f"Duplicate Jail '{name}' in uploaded file."
                )

            serializer = JailBulkUploadSerializer(
                data=data
            )

            serializer.is_valid(
                raise_exception=True
            )

            records.append(
                serializer.validated_data
            )

            uploaded_names.add(name)

        except Exception as exc:

            errors.append({
                "row": row_number,
                "error": str(exc)
            })

    # Do not save anything if any row has an error
    if errors:
        return {
            "message": "Bulk upload failed. No records were saved.",
            "created": 0,
            "failed": len(errors),
            "errors": errors,
        }

    # Save only when the complete file is valid
    created_count = save_records(
        Jail,
        records
    )

    return {
        "message": "Jail bulk upload completed successfully.",
        "created": created_count,
        "failed": 0,
        "errors": [],
    }

def bulk_upload_vegetables(uploaded_file):

    workbook = load_excel_file(uploaded_file)
    worksheet = get_active_sheet(workbook)

    headers = validate_headers(
        worksheet,
        VEGETABLE_HEADERS
    )

    records = []
    errors = []

    existing_items = set(
        VegetableMaster.objects.values_list(
            "item_name",
            flat=True
        )
    )

    uploaded_items = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if is_empty_row(row):
            continue

        data = row_to_dict(headers, row)

        try:
            data["is_active"] = normalize_boolean(
                data.get("is_active")
            )

            item_name = str(
                data.get("item_name") or ""
            ).strip()

            data["item_name"] = item_name

            if not item_name:
                raise ValueError(
                    "Item name is required."
                )

            if item_name in existing_items:
                raise ValueError(
                    f"Vegetable '{item_name}' already exists."
                )

            if item_name in uploaded_items:
                raise ValueError(
                    f"Duplicate vegetable '{item_name}' "
                    "in uploaded file."
                )

            serializer = VegetableBulkUploadSerializer(
                data=data
            )

            serializer.is_valid(
                raise_exception=True
            )

            records.append(
                serializer.validated_data
            )

            uploaded_items.add(item_name)

        except Exception as exc:

            errors.append({
                "row": row_number,
                "error": str(exc)
            })

    if errors:
        return {
            "message": "Bulk upload failed. No records were saved.",
            "created": 0,
            "failed": len(errors),
            "errors": errors,
        }

    created_count = save_records(
        VegetableMaster,
        records
    )

    return {
        "message": "VegetableMaster bulk upload completed successfully.",
        "created": created_count,
        "failed": 0,
        "errors": [],
    }

def bulk_upload_shgs(uploaded_file):

    workbook = load_excel_file(uploaded_file)
    worksheet = get_active_sheet(workbook)

    headers = validate_headers(
        worksheet,
        SHG_HEADERS
    )

    records = []
    errors = []

    # Build Jail lookup using Jail NAME
    # Example:
    # {
    #     "Central Jail": 1,
    #     "Ludhiana Jail": 2
    # }
    existing_jails = {
        jail.name.strip(): jail.pk
        for jail in Jail.objects.all()
    }

    # Existing SHGs to prevent duplicate database records
    existing_shgs = set(
        SHG.objects.values_list(
            "name",
            "jail_id"
        )
    )

    # Duplicate SHGs inside uploaded file
    uploaded_shgs = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if is_empty_row(row):
            continue

        data = row_to_dict(
            headers,
            row
        )

        try:

            # --------------------------------
            # Boolean conversion
            # --------------------------------
            data["is_active"] = normalize_boolean(
                data.get("is_active")
            )

            # --------------------------------
            # SHG name
            # --------------------------------
            name = str(
                data.get("name") or ""
            ).strip()

            if not name:
                raise ValueError(
                    "SHG name is required."
                )

            data["name"] = name

            # --------------------------------
            # Jail name from Excel
            # --------------------------------
            jail_name = str(
                data.get("jail") or ""
            ).strip()

            if not jail_name:
                raise ValueError(
                    "Jail name is required."
                )

            # --------------------------------
            # Find corresponding Jail PK
            # --------------------------------
            jail_pk = existing_jails.get(
                jail_name
            )

            if not jail_pk:
                raise ValueError(
                    f"Jail '{jail_name}' does not exist."
                )

            # IMPORTANT:
            # Serializer expects Jail PK,
            # not Jail object.
            data["jail"] = jail_pk

            # --------------------------------
            # Duplicate check in database
            # --------------------------------
            duplicate_key = (
                name,
                jail_pk
            )

            if duplicate_key in existing_shgs:
                raise ValueError(
                    f"SHG '{name}' already exists "
                    f"for Jail '{jail_name}'."
                )

            # --------------------------------
            # Duplicate check in uploaded file
            # --------------------------------
            if duplicate_key in uploaded_shgs:
                raise ValueError(
                    f"Duplicate SHG '{name}' "
                    f"for Jail '{jail_name}' "
                    "in uploaded file."
                )

            # --------------------------------
            # Serializer validation
            # --------------------------------
            serializer = SHGBulkUploadSerializer(
                data=data
            )

            serializer.is_valid(
                raise_exception=True
            )

            records.append(
                serializer.validated_data
            )

            uploaded_shgs.add(
                duplicate_key
            )

        except Exception as exc:

            errors.append({
                "row": row_number,
                "error": str(exc)
            })

    # ----------------------------------------
    # Atomic behaviour:
    # If ANY row fails, save NOTHING
    # ----------------------------------------
    if errors:

        return {
            "message": (
                "Bulk upload failed. "
                "No records were saved."
            ),
            "created": 0,
            "failed": len(errors),
            "errors": errors,
        }

    # ----------------------------------------
    # Save all records
    # ----------------------------------------
    created_count = save_records(
        SHG,
        records
    )

    return {
        "message": (
            "SHG bulk upload completed "
            "successfully."
        ),
        "created": created_count,
        "failed": 0,
        "errors": [],
    }


def save_records(model, records):

    if not records:
        return 0

    objects = [
        model(**record)
        for record in records
    ]

    with transaction.atomic():

        created_objects = model.objects.bulk_create(
            objects
        )

    return len(created_objects)
